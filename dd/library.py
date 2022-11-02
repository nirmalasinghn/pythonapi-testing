import json
import jsonpath
import requests
import openpyxl

class chris:
    def __init__(self,FileName,SheetName):
        global vk
        global sh
        vk=openpyxl.load_workbook('C:\\Users\\akash\\OneDrive\\Desktop\\code\\excel1.xlsx')
        sh=vk['Sheet1']
        rows=sh.max_row
    def fetch_rows(self):
        rows=sh.max_row
        return rows

    def fetch_column(self):
        column=sh.max_column
        return column

    def fetch_keynames(self):
        c=sh.max_column
        li=[]
        for i in range(1,c+1):
            cell=sh.cell(row=1,column=i)
            li.insert(i-1,cell.value)
        return li   

    def update_req(self,rownum,jsonreq,keylist):
        c=sh.max_column
        for i in range(1,c+1):
            cell=sh.cell(row=rownum,column=i)
            jsonreq[keylist[i-1]]=cell.value
        return jsonreq