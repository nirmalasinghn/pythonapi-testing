import requests
import json 
import jsonpath
import openpyxl


def test_p():
    #api
    url='https://thetestingworldapi.com/api/studentsDetails'
    f=open('C:\\Users\\akash\\OneDrive\\Desktop\\code\\kobe.json')
    json_req=json.loads(f.read())
   
    #excel code
    vk=openpyxl.load_workbook('C:\\Users\\akash\\OneDrive\\Desktop\\code\\excel1.xlsx')
    sh=vk['Sheet1']
    rows=sh.max_row
    
    for i in range(2,rows+1):
        cell_firstname=sh.cell(row=i, column=1)
        cell_midname=sh.cell(row=i, column=2)
        cell_lastname=sh.cell(row=i, column=3)
        cell_dateofbirth=sh.cell(row=i, column=4)
    
        json_req['first_name']=cell_firstname.value
        json_req['middle_name']=cell_midname.value
        json_req['last_name']=cell_lastname.value
        json_req['date_of_birth']=cell_dateofbirth.value
        
        res=requests.post(url,json_req)
        
        print(res.text)
        print(res.status_code)
        assert res.status_code ==201